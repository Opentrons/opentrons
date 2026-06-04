# -*- coding: utf-8 -*-
from __future__ import annotations

import math
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from reportlab.graphics.shapes import Circle, Drawing, Group, Line, Rect, String
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.platypus import (
    BaseDocTemplate,
    Frame,
    KeepTogether,
    PageBreak,
    PageTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
)


ROOT = Path("/Users/yew/opentrons")
XLSX = Path("/tmp/lifetime-report-96-p1000h-latest.xlsx")
OUT = ROOT / "aging_test_report_96_p1000h_per_pipette_with_charts.pdf"

SERIES_COLORS = [
    colors.HexColor("#1F77B4"),
    colors.HexColor("#D62728"),
    colors.HexColor("#2CA02C"),
    colors.HexColor("#9467BD"),
    colors.HexColor("#FF7F0E"),
]


def is_num(v) -> bool:
    return isinstance(v, (int, float)) and not isinstance(v, bool) and not math.isnan(v)


def fmt(v, digits=3) -> str:
    if is_num(v):
        return f"{v:.{digits}f}"
    if v in (None, ""):
        return "-"
    return str(v)


def fmt_cycle(v) -> str:
    return f"{int(v):,}" if is_num(v) else "-"


def fmt_life(v) -> str:
    if is_num(v):
        return f"{v * 100:.1f}%"
    return str(v) if v not in (None, "") else "-"


def clean(v) -> str:
    if v is None:
        return ""
    return str(v).replace("\n", " ").strip()


def pct_name(pipette: str) -> str:
    return pipette.replace("P1KHV3620", "")


def extract_data():
    wb = load_workbook(XLSX, data_only=True, read_only=True)
    ws = wb["RD Lifetime test-96ch P1KH"]
    rows = list(ws.iter_rows(values_only=True))

    starts = []
    for i, row in enumerate(rows, start=1):
        if len(row) > 1 and row[1] and str(row[1]).startswith("P1K"):
            starts.append(i)

    samples = []
    for idx, start in enumerate(starts):
        end = starts[idx + 1] - 1 if idx + 1 < len(starts) else 92
        block = rows[start - 1 : end]
        first = block[0]
        pipette = first[1]

        numeric = []
        all_cycle_records = []
        notes = []
        fails = []
        pass_checks = []

        for excel_row, row in enumerate(block, start=start):
            cycles = row[4] if len(row) > 4 else None
            life = row[7] if len(row) > 7 else None
            cv = row[8] if len(row) > 8 else None
            plates = list(row[9:14]) if len(row) > 13 else []
            worst = row[14] if len(row) > 14 else None
            result_cells = list(row[8:15]) if len(row) > 14 else []
            note = row[17] if len(row) > 17 else None
            failed_channels = row[19] if len(row) > 19 else None

            if is_num(cycles):
                record = {
                    "excel_row": excel_row,
                    "cycles": cycles,
                    "life": life,
                    "cv": cv,
                    "plates": plates,
                    "worst": worst,
                    "note": clean(note),
                    "failed_channels": clean(failed_channels),
                    "result_cells": result_cells,
                    "numeric_plates": sum(1 for x in plates if is_num(x)),
                }
                all_cycle_records.append(record)
                if is_num(cv):
                    numeric.append(record)

                result_words = [str(v).strip().lower() for v in result_cells if v not in (None, "")]
                if "fail" in result_words:
                    fails.append(record)
                elif result_words and all(v == "pass" for v in result_words):
                    pass_checks.append(record)

            if note not in (None, "") or failed_channels not in (None, ""):
                notes.append(
                    {
                        "cycles": cycles,
                        "life": life,
                        "note": clean(note),
                        "failed_channels": clean(failed_channels),
                    }
                )

        latest = max(numeric, key=lambda r: r["cycles"])
        max_cv = max(numeric, key=lambda r: r["cv"])
        max_worst = max([r for r in numeric if is_num(r["worst"])], key=lambda r: r["worst"])
        avg_cv = sum(r["cv"] for r in numeric) / len(numeric)
        last_three = sorted(numeric, key=lambda r: r["cycles"])[-3:]
        last_three_avg = sum(r["cv"] for r in last_three) / len(last_three)

        samples.append(
            {
                "pipette": pipette,
                "status": clean(first[0]),
                "numeric": sorted(numeric, key=lambda r: r["cycles"]),
                "records": sorted(all_cycle_records, key=lambda r: r["cycles"]),
                "latest": latest,
                "max_cv": max_cv,
                "max_worst": max_worst,
                "avg_cv": avg_cv,
                "last_three_avg": last_three_avg,
                "notes": notes,
                "fails": fails,
                "pass_checks": pass_checks,
            }
        )

    plan = wb["Test Plan"]
    plan_rows = list(plan.iter_rows(values_only=True))
    plan_notes = []
    for row in plan_rows:
        vals = [clean(v) for v in row if v not in (None, "")]
        joined = "；".join(vals)
        if "继续跑到fail" in joined or "取消每 65,000" in joined or "2,000,000" in joined:
            plan_notes.append(joined)

    return {"samples": samples, "plan_notes": plan_notes}


def styles():
    pdfmetrics.registerFont(UnicodeCIDFont("STSong-Light"))
    base = getSampleStyleSheet()

    def make(name, **kw):
        parent = kw.pop("parent", base["Normal"])
        return ParagraphStyle(name, parent=parent, fontName="STSong-Light", wordWrap="CJK", **kw)

    return {
        "title": make("title", fontSize=22, leading=28, alignment=TA_CENTER, spaceAfter=6),
        "subtitle": make("subtitle", fontSize=9, leading=12, alignment=TA_CENTER, textColor=colors.HexColor("#5A5A5A")),
        "h1": make("h1", fontSize=14, leading=18, textColor=colors.HexColor("#1F4E79"), spaceBefore=8, spaceAfter=5),
        "h2": make("h2", fontSize=11.5, leading=15, textColor=colors.HexColor("#333333"), spaceBefore=5, spaceAfter=4),
        "body": make("body", fontSize=9, leading=12.8, spaceAfter=3),
        "small": make("small", fontSize=7.5, leading=9.5),
        "table": make("table", fontSize=7.2, leading=9.2),
        "table_center": make("table_center", fontSize=7.2, leading=9.2, alignment=TA_CENTER),
        "note": make("note", fontSize=8.7, leading=12, leftIndent=4, rightIndent=4),
    }


def para(text, style):
    return Paragraph(str(text).replace("\n", "<br/>"), style)


def make_table(rows, widths, header=True, font_size=7.1):
    table = Table(rows, colWidths=widths, repeatRows=1 if header else 0, hAlign="LEFT")
    commands = [
        ("FONTNAME", (0, 0), (-1, -1), "STSong-Light"),
        ("FONTSIZE", (0, 0), (-1, -1), font_size),
        ("LEADING", (0, 0), (-1, -1), font_size + 2),
        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#BECBD6")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 3.5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3.5),
        ("LEFTPADDING", (0, 0), (-1, -1), 3.5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3.5),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FBFD")]),
    ]
    if header:
        commands.extend(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#D9EAF7")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#163B59")),
            ]
        )
    table.setStyle(TableStyle(commands))
    return table


def chart_axes(d, x0, y0, w, h, y_max, title, y_label, x_ticks):
    d.add(String(x0 + w / 2, y0 + h + 24, title, fontName="STSong-Light", fontSize=10, textAnchor="middle", fillColor=colors.HexColor("#1F4E79")))
    d.add(String(x0 - 28, y0 + h / 2, y_label, fontName="STSong-Light", fontSize=7, textAnchor="middle", fillColor=colors.HexColor("#555555")))
    d.add(Line(x0, y0, x0 + w, y0, strokeColor=colors.HexColor("#666666"), strokeWidth=0.8))
    d.add(Line(x0, y0, x0, y0 + h, strokeColor=colors.HexColor("#666666"), strokeWidth=0.8))

    for i in range(5):
        value = y_max * i / 4
        y = y0 + h * i / 4
        d.add(Line(x0, y, x0 + w, y, strokeColor=colors.HexColor("#E4E9EE"), strokeWidth=0.45))
        d.add(String(x0 - 5, y - 2, f"{value:.1f}", fontName="STSong-Light", fontSize=6.5, textAnchor="end", fillColor=colors.HexColor("#555555")))

    for value, label in x_ticks:
        x = x0 + w * value / 1020000
        d.add(Line(x, y0 - 3, x, y0, strokeColor=colors.HexColor("#666666"), strokeWidth=0.5))
        d.add(String(x, y0 - 13, label, fontName="STSong-Light", fontSize=6.2, textAnchor="middle", fillColor=colors.HexColor("#555555")))


def line_chart(samples, metric, title, y_label, width=245 * mm, height=86 * mm):
    d = Drawing(width, height)
    margin_left, margin_right, margin_top, margin_bottom = 42, 20, 35, 24
    x0, y0 = margin_left, margin_bottom
    w, h = width - margin_left - margin_right, height - margin_top - margin_bottom
    y_max = max((r[metric] for s in samples for r in s["numeric"] if is_num(r.get(metric))), default=1)
    y_max = max(1, math.ceil(y_max * 1.12))
    x_ticks = [(0, "0"), (260000, "260k"), (520000, "520k"), (820000, "820k"), (1020000, "1020k")]
    chart_axes(d, x0, y0, w, h, y_max, title, y_label, x_ticks)

    for i, sample in enumerate(samples):
        color = SERIES_COLORS[i % len(SERIES_COLORS)]
        points = []
        for r in sample["numeric"]:
            v = r.get(metric)
            if not is_num(v):
                continue
            x = x0 + w * r["cycles"] / 1020000
            y = y0 + h * v / y_max
            points.append((x, y))
        for (x1, y1), (x2, y2) in zip(points, points[1:]):
            d.add(Line(x1, y1, x2, y2, strokeColor=color, strokeWidth=1.4))
        for x, y in points:
            d.add(Circle(x, y, 2.1, fillColor=color, strokeColor=color))

        lx = x0 + 5 + (i % 3) * 118
        ly = y0 + h + 7 - (i // 3) * 10
        d.add(Rect(lx, ly - 3, 7, 7, fillColor=color, strokeColor=color))
        d.add(String(lx + 10, ly - 2, pct_name(sample["pipette"]), fontName="STSong-Light", fontSize=6.6, fillColor=colors.HexColor("#333333")))
    return d


def bar_chart(items, title, y_label, width=118 * mm, height=76 * mm):
    d = Drawing(width, height)
    x0, y0 = 32, 22
    w, h = width - 48, height - 50
    max_v = max(v for _, v in items) if items else 1
    y_max = max(1, math.ceil(max_v * 1.18))
    chart_axes(d, x0, y0, w, h, y_max, title, y_label, [(0, "0"), (1020000, "")])
    bar_gap = 5
    bar_w = (w - bar_gap * (len(items) - 1)) / len(items)
    for i, (label, value) in enumerate(items):
        color = SERIES_COLORS[i % len(SERIES_COLORS)]
        x = x0 + i * (bar_w + bar_gap)
        bh = h * value / y_max
        d.add(Rect(x, y0, bar_w, bh, fillColor=color, strokeColor=color))
        d.add(String(x + bar_w / 2, y0 - 11, pct_name(label), fontName="STSong-Light", fontSize=5.8, textAnchor="middle"))
        d.add(String(x + bar_w / 2, y0 + bh + 4, f"{value:.2f}", fontName="STSong-Light", fontSize=6, textAnchor="middle"))
    return d


def small_dual_chart(sample, width=116 * mm, height=54 * mm):
    d = Drawing(width, height)
    x0, y0 = 30, 15
    w, h = width - 43, height - 28
    y_max = max(
        max(r["cv"] for r in sample["numeric"] if is_num(r["cv"])),
        max(r["worst"] for r in sample["numeric"] if is_num(r["worst"])),
    )
    y_max = max(1, math.ceil(y_max * 1.15))
    chart_axes(d, x0, y0, w, h, y_max, "CV & worst-channel trend", "CV", [(0, "0"), (520000, "520k"), (1020000, "1020k")])
    for metric, color, label, off in [
        ("cv", colors.HexColor("#1F77B4"), "CV", 0),
        ("worst", colors.HexColor("#D62728"), "Worst", 32),
    ]:
        pts = []
        for r in sample["numeric"]:
            if not is_num(r.get(metric)):
                continue
            pts.append((x0 + w * r["cycles"] / 1020000, y0 + h * r[metric] / y_max))
        for a, b in zip(pts, pts[1:]):
            d.add(Line(a[0], a[1], b[0], b[1], strokeColor=color, strokeWidth=1.1))
        for x, y in pts:
            d.add(Circle(x, y, 1.8, fillColor=color, strokeColor=color))
        d.add(Rect(x0 + off, y0 + h + 3, 6, 6, fillColor=color, strokeColor=color))
        d.add(String(x0 + off + 8, y0 + h + 3, label, fontName="STSong-Light", fontSize=6))
    return d


def trend_summary(sample):
    p = sample["pipette"]
    latest = sample["latest"]
    max_cv = sample["max_cv"]
    max_worst = sample["max_worst"]
    notes = [n["note"] for n in sample["notes"] if n["note"]]
    fail = sample["fails"]

    if "0624A05" in p:
        return (
            "最新节点为全组最大风险：1020k 时 CV=6.839，plate1-3 分别为 6.743/6.944/6.898，"
            "说明不是单盘偶发，而是三盘同步偏高。建议优先复测光学，并确认三盘口径是否会放大 CV。"
        )
    if "0624A04" in p:
        return (
            "早期 195k-260k CV 抬升，520k 再次出现较高 CV 和 worst-channel CV 峰值；"
            "同时备注记录 F1/A5 LOW %D、520k 取针管报错/漏油/传感器报错。虽最新 worst-channel CV 回落，"
            "但历史事件需要作为老化异常项关闭。"
        )
    if "1217A05" in p:
        return (
            "175%/455k 节点主表有明确 fail，备注为 C5/C7 两个通道吸水少；"
            "520k 出现 CV 峰值 6.195，之后最新 CV 回落到 1.487。该样机不能只按最新值判定，应保留中途失败记录。"
        )
    if "0311A04" in p:
        return (
            "整体最稳定。除 0 cycle 初始 CV=1.957 外，后续节点 CV 基本低于 1.2，"
            "1020k 最新 CV=0.677、worst-channel CV=1.052，主表未记录非 Pressure 显式异常。"
        )
    if "0624A01" in p:
        return (
            "早期波动最大，195k 节点 worst-channel CV=13.397，并记录 H1/G12 LOW %D；"
            "520k CV 达 8.406。620k 后逐步回落，1020k 最新 CV=1.633。建议保留历史异常并继续观察。"
        )
    return f"最新 CV={fmt(latest['cv'])}，最大 CV={fmt(max_cv['cv'])}，最大 worst-channel CV={fmt(max_worst['worst'])}。"


def build_pdf(data):
    st = styles()
    page_size = landscape(A4)
    doc = BaseDocTemplate(
        str(OUT),
        pagesize=page_size,
        leftMargin=15 * mm,
        rightMargin=15 * mm,
        topMargin=13 * mm,
        bottomMargin=13 * mm,
        title="P1KH 96-Channel P1000H Aging Test Report",
    )
    frame = Frame(doc.leftMargin, doc.bottomMargin, doc.width, doc.height, id="normal")

    def footer(canvas, document):
        canvas.saveState()
        canvas.setFont("STSong-Light", 7)
        canvas.setFillColor(colors.HexColor("#666666"))
        canvas.drawString(doc.leftMargin, 8 * mm, "P1KH 96-Channel P1000H Aging Test Report | Pressure excluded")
        canvas.drawRightString(page_size[0] - doc.rightMargin, 8 * mm, f"Page {document.page}")
        canvas.restoreState()

    doc.addPageTemplates([PageTemplate(id="main", frames=[frame], onPage=footer)])
    samples = data["samples"]
    story = []

    story.append(para("P1KH 96-Channel P1000H 老化测试分析报告", st["title"]))
    story.append(para("报告日期：2026-06-03 | 数据源：Life Time Report -96 P1000H | 分析口径：不分析 Pressure 测试结果", st["subtitle"]))
    story.append(Spacer(1, 5))

    latest_cycles = max(s["latest"]["cycles"] for s in samples)
    latest_life = max(s["latest"]["life"] for s in samples if is_num(s["latest"]["life"]))
    issue_count = sum(len(s["fails"]) + len(s["notes"]) for s in samples)
    callout = (
        f"<b>最终结论：</b>5 台 P1KH 96-channel P1000H 移液器均已完成老化循环到 "
        f"{fmt_cycle(latest_cycles)} cycles（约 {fmt_life(latest_life)}）。"
        f"按非 Pressure 数据，不能简单写为全组无异常通过：主表共有 {issue_count} 条 fail/备注类异常记录，"
        f"其中 `P1KHV3620241217A05` 有 175% 显式 fail，`P1KHV3620250624A05` 最新 1020k 光学 CV 明显升高，"
        f"`P1KHV3620250624A04` 与 `P1KHV3620250624A01` 有历史 LOW %D/高 CV 节点。"
        f"`P1KHV3620250311A04` 是本轮老化中表现最稳定的样机。"
    )
    callout_table = Table([[para(callout, st["note"])]], colWidths=[doc.width])
    callout_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#FFF7E6")),
                ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#E2B76B")),
                ("LEFTPADDING", (0, 0), (-1, -1), 8),
                ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                ("TOPPADDING", (0, 0), (-1, -1), 7),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
            ]
        )
    )
    story.append(callout_table)

    story.append(para("1. 测试范围与分析口径", st["h1"]))
    for item in [
        "本报告是老化测试分析报告，分析对象为 5 台 production pipette，测试类型为 1000 uL cycles。",
        "纳入数据：主寿命表中的 cycles、% Life、CV across all dispenses、plate1-plate5、CV of the worst channel、PASS/FAIL、Notes、Fail Image、Failed Channels，以及 Test Plan 的老化计划备注。",
        "明确排除：所有 Pressure 工作表、Pressure Test link、外链 Pressure Summary、压力阈值、压力 leak-rate / holding / insert 结果。",
        "最新状态：5 台样机主表状态均为 Finish，最新记录均为 1,020,000 cycles / 392.3% life。",
    ]:
        story.append(para("- " + item, st["body"]))

    story.append(para("2. 整体趋势图", st["h1"]))
    story.append(line_chart(samples, "cv", "Overall CV trend", "CV"))
    story.append(Spacer(1, 4))
    story.append(line_chart(samples, "worst", "Worst-channel trend", "Worst CV"))
    story.append(PageBreak())

    story.append(para("3. 整体对比图与汇总表", st["h1"]))
    bar_items_latest = [(s["pipette"], s["latest"]["cv"]) for s in samples]
    bar_items_max = [(s["pipette"], s["max_cv"]["cv"]) for s in samples]
    charts = Table(
        [[bar_chart(bar_items_latest, "Latest CV @1020k", "CV"), bar_chart(bar_items_max, "Max historical CV", "CV")]],
        colWidths=[130 * mm, 130 * mm],
    )
    charts.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP")]))
    story.append(charts)
    story.append(Spacer(1, 6))

    summary = [[para(x, st["table_center"]) for x in ["Pipette", "最新 CV", "最大 CV", "最大 worst CV", "显式 fail", "备注数", "老化风险", "一句话判断"]]]
    for s in samples:
        risk = "高" if "0624A05" in s["pipette"] else "中高" if s["fails"] or "0624A04" in s["pipette"] else "中" if "0624A01" in s["pipette"] else "低"
        one_line = trend_summary(s)
        summary.append(
            [
                para(s["pipette"], st["table"]),
                para(fmt(s["latest"]["cv"]), st["table_center"]),
                para(f"{fmt(s['max_cv']['cv'])}<br/>{fmt_cycle(s['max_cv']['cycles'])}", st["table_center"]),
                para(f"{fmt(s['max_worst']['worst'])}<br/>{fmt_cycle(s['max_worst']['cycles'])}", st["table_center"]),
                para(str(len(s["fails"])), st["table_center"]),
                para(str(len(s["notes"])), st["table_center"]),
                para(risk, st["table_center"]),
                para(one_line, st["table"]),
            ]
        )
    story.append(make_table(summary, [30 * mm, 18 * mm, 24 * mm, 28 * mm, 18 * mm, 16 * mm, 18 * mm, 108 * mm]))
    story.append(PageBreak())

    story.append(para("4. 每台移液器老化分析", st["h1"]))
    for idx, s in enumerate(samples, start=1):
        story.append(KeepTogether([para(f"4.{idx} {s['pipette']}", st["h2"]), small_dual_chart(s)]))
        latest = s["latest"]
        max_cv = s["max_cv"]
        max_worst = s["max_worst"]
        facts = (
            f"状态：{s['status']}；最新：{fmt_cycle(latest['cycles'])} cycles / {fmt_life(latest['life'])}；"
            f"最新 CV={fmt(latest['cv'])}，最新 worst-channel CV={fmt(latest['worst'])}；"
            f"历史最大 CV={fmt(max_cv['cv'])}（{fmt_cycle(max_cv['cycles'])}, {fmt_life(max_cv['life'])}），"
            f"历史最大 worst-channel CV={fmt(max_worst['worst'])}（{fmt_cycle(max_worst['cycles'])}, {fmt_life(max_worst['life'])}）。"
        )
        story.append(para(facts, st["body"]))
        story.append(para("<b>分析：</b>" + trend_summary(s), st["body"]))

        rows = [[para(x, st["table_center"]) for x in ["cycles", "% life", "CV", "worst CV", "plate 数据", "备注/结果"]]]
        interesting = []
        for rec in s["numeric"]:
            if rec in (s["latest"], s["max_cv"], s["max_worst"]) or rec["note"]:
                interesting.append(rec)
        for rec in s["fails"]:
            if rec not in interesting:
                interesting.append(rec)
        interesting = sorted({r["excel_row"]: r for r in interesting}.values(), key=lambda r: r["cycles"])
        for rec in interesting:
            plate_text = " / ".join(fmt(v) for v in rec["plates"])
            result_words = [clean(v) for v in rec["result_cells"] if v not in (None, "")]
            note = rec["note"] or ("; ".join(result_words) if any(str(v).lower() == "fail" for v in result_words) else "")
            rows.append(
                [
                    para(fmt_cycle(rec["cycles"]), st["table_center"]),
                    para(fmt_life(rec["life"]), st["table_center"]),
                    para(fmt(rec["cv"]), st["table_center"]),
                    para(fmt(rec["worst"]), st["table_center"]),
                    para(plate_text, st["table"]),
                    para(note or "-", st["table"]),
                ]
            )
        story.append(make_table(rows, [23 * mm, 19 * mm, 17 * mm, 20 * mm, 80 * mm, 101 * mm]))
        story.append(Spacer(1, 6))

    story.append(PageBreak())
    story.append(para("5. 异常清单（非 Pressure）", st["h1"]))
    event_rows = [[para(x, st["table_center"]) for x in ["Pipette", "节点", "类型", "内容", "处理建议"]]]
    for s in samples:
        for fail in s["fails"]:
            event_rows.append(
                [
                    para(s["pipette"], st["table"]),
                    para(f"{fmt_cycle(fail['cycles'])}<br/>{fmt_life(fail['life'])}", st["table_center"]),
                    para("显式 fail", st["table_center"]),
                    para(fail["note"] or "结果单元格含 fail。", st["table"]),
                    para("需作为老化异常闭环，不能仅用最新 CV 覆盖。", st["table"]),
                ]
            )
        for note in s["notes"]:
            event_rows.append(
                [
                    para(s["pipette"], st["table"]),
                    para(f"{fmt_cycle(note['cycles'])}<br/>{fmt_life(note['life'])}", st["table_center"]),
                    para("备注", st["table_center"]),
                    para(note["note"] or note["failed_channels"], st["table"]),
                    para("复核原始图片/光学数据，确认是否复现及是否已关闭。", st["table"]),
                ]
            )
    story.append(make_table(event_rows, [34 * mm, 28 * mm, 22 * mm, 112 * mm, 64 * mm]))

    story.append(para("6. 最终总结与建议", st["h1"]))
    final_points = [
        "<b>完成度：</b>5 台样机均完成 1,020,000 cycles，达到约 392.3% life，主表状态为 Finish。",
        "<b>最佳样机：</b>P1KHV3620250311A04，1020k CV=0.677，worst-channel CV=1.052，未记录非 Pressure 显式异常。",
        "<b>最高风险：</b>P1KHV3620250624A05，最新 1020k CV=6.839，且 plate1-3 同步偏高，应优先复测。",
        "<b>中途失败样机：</b>P1KHV3620241217A05 在 455k / 175% 记录 fail，备注 C5/C7 吸水少，应保留为老化异常事件。",
        "<b>历史异常样机：</b>P1KHV3620250624A04 有 LOW %D 与漏油/传感器报错记录；P1KHV3620250624A01 有 LOW %D 与历史高 CV，但最新均有回落。",
        "<b>报告限制：</b>本报告未分析 Pressure 测试结果；若最终质量判定需要 pressure 数据，应另出 Pressure 专项报告或合并评审。",
    ]
    for item in final_points:
        story.append(para("- " + item, st["body"]))

    doc.build(story)


if __name__ == "__main__":
    build_pdf(extract_data())
    print(OUT)
