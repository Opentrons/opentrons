# -*- coding: utf-8 -*-
from __future__ import annotations

import math
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    BaseDocTemplate,
    Frame,
    KeepTogether,
    PageTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
)
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont


ROOT = Path("/Users/yew/opentrons")
XLSX = Path("/tmp/lifetime-report-96-p1000h-latest.xlsx")
OUT = ROOT / "lifetime_test_report_96_p1000h_no_pressure.pdf"


def is_num(value) -> bool:
    return isinstance(value, (int, float)) and not isinstance(value, bool) and not math.isnan(value)


def fmt_num(value, digits=3) -> str:
    if not is_num(value):
        return str(value) if value not in (None, "") else "-"
    return f"{value:.{digits}f}"


def fmt_cycles(value) -> str:
    if not is_num(value):
        return "-"
    return f"{int(value):,}"


def fmt_life(value) -> str:
    if is_num(value):
        return f"{value * 100:.1f}%"
    return str(value) if value not in (None, "") else "-"


def cell_text(value) -> str:
    if value is None:
        return ""
    return str(value).replace("\n", " ").strip()


def to_date(value):
    if value in (None, ""):
        return None
    if hasattr(value, "date"):
        return value.date()
    if isinstance(value, str):
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
            try:
                return datetime.strptime(value.strip(), fmt).date()
            except ValueError:
                pass
    return None


def extract_data():
    wb = load_workbook(XLSX, data_only=True, read_only=True)
    ws = wb["RD Lifetime test-96ch P1KH"]

    rows = list(ws.iter_rows(values_only=True))
    starts = []
    for i, row in enumerate(rows, start=1):
        if len(row) > 1 and row[1] and str(row[1]).startswith("P1K"):
            starts.append(i)

    samples = []
    for idx, start in enumerate(starts):
        end = starts[idx + 1] - 1 if idx + 1 < len(starts) else 92
        block = rows[start - 1 : end]
        first = block[0]
        pipette = first[1]
        status = cell_text(first[0])

        cycle_rows = []
        numeric_rows = []
        notes = []
        explicit_fails = []
        text_pass_rows = []
        for offset, row in enumerate(block, start=start):
            cycles = row[4] if len(row) > 4 else None
            life = row[7] if len(row) > 7 else None
            cv = row[8] if len(row) > 8 else None
            plates = list(row[9:14]) if len(row) > 13 else []
            worst = row[14] if len(row) > 14 else None
            link = row[15] if len(row) > 15 else None
            note = row[17] if len(row) > 17 else None
            failed_channels = row[19] if len(row) > 19 else None
            result_cells = list(row[8:15]) if len(row) > 14 else []

            if is_num(cycles):
                record = {
                    "row": offset,
                    "cycles": cycles,
                    "life": life,
                    "cv": cv,
                    "plates": plates,
                    "worst": worst,
                    "link": link,
                    "note": note,
                    "failed_channels": failed_channels,
                    "result_cells": result_cells,
                }
                cycle_rows.append(record)
                if is_num(cv):
                    numeric_rows.append(record)

            if note not in (None, "") or failed_channels not in (None, ""):
                notes.append(
                    {
                        "row": offset,
                        "cycles": cycles,
                        "life": life,
                        "note": cell_text(note),
                        "failed_channels": cell_text(failed_channels),
                    }
                )

            lowered = [str(v).strip().lower() for v in result_cells if v not in (None, "")]
            if "fail" in lowered:
                explicit_fails.append(
                    {
                        "row": offset,
                        "cycles": cycles,
                        "life": life,
                        "note": cell_text(note),
                        "cells": result_cells,
                    }
                )
            elif lowered and all(v == "pass" for v in lowered):
                text_pass_rows.append({"row": offset, "cycles": cycles, "life": life})

        latest = max(cycle_rows, key=lambda r: r["cycles"])
        latest_numeric = max(numeric_rows, key=lambda r: r["cycles"])
        max_cv = max(numeric_rows, key=lambda r: r["cv"])
        max_worst = max(
            [r for r in numeric_rows if is_num(r["worst"])],
            key=lambda r: r["worst"],
        )
        latest_plate_count = sum(1 for v in latest_numeric["plates"] if is_num(v))

        samples.append(
            {
                "pipette": pipette,
                "status": status,
                "cycle_rows": cycle_rows,
                "numeric_rows": numeric_rows,
                "latest": latest,
                "latest_numeric": latest_numeric,
                "max_cv": max_cv,
                "max_worst": max_worst,
                "latest_plate_count": latest_plate_count,
                "notes": notes,
                "explicit_fails": explicit_fails,
                "text_pass_rows": text_pass_rows,
            }
        )

    plan = wb["Test Plan"]
    plan_rows = list(plan.iter_rows(values_only=True))
    start_dates = []
    finish_392_dates = []
    plan_notes = []
    for row in plan_rows[1:6]:
        if row[2]:
            start_dates.append(row[2])
        if len(row) > 29 and row[29]:
            finish_392_dates.append(row[29])
    for row in plan_rows:
        vals = [cell_text(v) for v in row if v not in (None, "")]
        if any("继续跑到fail" in v or "取消每 65,000" in v or "总运行次数到 2,000,000" in v for v in vals):
            plan_notes.append("；".join(vals))

    return {
        "samples": samples,
        "start_date": min([d for d in (to_date(v) for v in start_dates) if d], default=None),
        "finish_392": max([d for d in (to_date(v) for v in finish_392_dates) if d], default=None),
        "plan_notes": plan_notes,
    }


def setup_styles():
    pdfmetrics.registerFont(UnicodeCIDFont("STSong-Light"))
    styles = getSampleStyleSheet()

    def s(name, **kwargs):
        base = kwargs.pop("parent", styles["Normal"])
        return ParagraphStyle(name, parent=base, fontName="STSong-Light", wordWrap="CJK", **kwargs)

    return {
        "title": s("TitleCN", fontSize=22, leading=27, alignment=TA_CENTER, spaceAfter=8),
        "subtitle": s("SubtitleCN", fontSize=9.5, leading=13, alignment=TA_CENTER, textColor=colors.HexColor("#555555")),
        "h1": s("H1CN", fontSize=13.5, leading=18, spaceBefore=9, spaceAfter=5, textColor=colors.HexColor("#1F4E79")),
        "h2": s("H2CN", fontSize=11.5, leading=15, spaceBefore=7, spaceAfter=4, textColor=colors.HexColor("#333333")),
        "body": s("BodyCN", fontSize=9.2, leading=13, spaceAfter=4),
        "small": s("SmallCN", fontSize=8, leading=10.5),
        "table": s("TableCN", fontSize=7.4, leading=9.3),
        "table_center": s("TableCenterCN", fontSize=7.3, leading=9.2, alignment=TA_CENTER),
        "right": s("RightCN", fontSize=7.4, leading=9.3, alignment=TA_RIGHT),
        "callout": s("CalloutCN", fontSize=9.2, leading=13, leftIndent=5, rightIndent=5, spaceAfter=4),
    }


def p(text, style):
    return Paragraph(str(text).replace("\n", "<br/>"), style)


def make_table(data, col_widths, header_rows=1, font_size=7.2):
    table = Table(data, colWidths=col_widths, repeatRows=header_rows, hAlign="LEFT")
    table.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (-1, -1), "STSong-Light"),
                ("FONTSIZE", (0, 0), (-1, -1), font_size),
                ("LEADING", (0, 0), (-1, -1), font_size + 2),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#D9EAF7")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#163B59")),
                ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#B9C7D3")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F7FAFC")]),
            ]
        )
    )
    return table


def classify(sample):
    pipette = sample["pipette"]
    latest_cv = sample["latest_numeric"]["cv"]
    max_cv = sample["max_cv"]["cv"]
    has_fail = bool(sample["explicit_fails"])
    notes = " ".join(n["note"] for n in sample["notes"])

    if pipette.endswith("A05") and "0624A05" in pipette:
        return "高", "最新 1020k 光学 CV 达 6.839，plate1-3 同时偏高，需优先复测。"
    if "0624A04" in pipette:
        return "中高", "历史低 %D 备注和 520k 漏油/传感器报错，最新 CV 仍偏高。"
    if has_fail:
        return "中高", "175% 节点有明确 fail，备注 C5/C7 吸水少。"
    if "0624A01" in pipette:
        return "中", "历史 CV 峰值较高并有 H1/G12 low %D，最新已回落但需跟踪。"
    if latest_cv <= 1 and max_cv < 2.5 and not notes:
        return "低", "主表光学/CV 表现最稳定，未记录显式非 pressure 异常。"
    return "中", "存在历史波动，建议按常规复核。"


def build_pdf(data):
    styles = setup_styles()
    page_size = landscape(A4)
    doc = BaseDocTemplate(
        str(OUT),
        pagesize=page_size,
        rightMargin=16 * mm,
        leftMargin=16 * mm,
        topMargin=15 * mm,
        bottomMargin=14 * mm,
        title="P1KH 96-Channel P1000H Lifetime Test Report - No Pressure",
    )
    frame = Frame(doc.leftMargin, doc.bottomMargin, doc.width, doc.height, id="normal")

    def footer(canvas, document):
        canvas.saveState()
        canvas.setFont("STSong-Light", 7)
        canvas.setFillColor(colors.HexColor("#666666"))
        canvas.drawString(doc.leftMargin, 9 * mm, "P1KH 96-Channel P1000H Lifetime Test Report | Pressure results excluded")
        canvas.drawRightString(page_size[0] - doc.rightMargin, 9 * mm, f"Page {document.page}")
        canvas.restoreState()

    doc.addPageTemplates([PageTemplate(id="main", frames=[frame], onPage=footer)])

    story = []
    today = datetime(2026, 6, 3).strftime("%Y-%m-%d")
    story.append(p("P1KH 96-Channel P1000H Lifetime Test Report", styles["title"]))
    story.append(
        p(
            f"报告日期：{today} | 数据源：Life Time Report -96 P1000H | 分析口径：不纳入 Pressure 测试结果",
            styles["subtitle"],
        )
    )
    story.append(Spacer(1, 6))

    samples = data["samples"]
    total = len(samples)
    finished = sum(1 for s in samples if s["status"].lower() == "finish")
    latest_cycles = max(s["latest"]["cycles"] for s in samples)
    latest_life = max(s["latest"]["life"] for s in samples if is_num(s["latest"]["life"]))
    explicit_fail_count = sum(len(s["explicit_fails"]) for s in samples)

    callout = (
        f"<b>结论：</b>按非 Pressure 口径，5 台样机均已达到 {fmt_cycles(latest_cycles)} cycles "
        f"（约 {fmt_life(latest_life)} life），主表状态均为 Finish。"
        f"但不能写成“完全无异常通过”：主表仍有 1 个明确 fail 节点、若干 LOW %D/吸水少备注，"
        f"以及 A05 最新光学 CV 明显升高。建议将结论定义为“寿命循环目标完成；非 Pressure 数据存在需复核项”。"
    )
    callout_table = Table([[p(callout, styles["callout"])]], colWidths=[doc.width])
    callout_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#FFF7E6")),
                ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#E2B76B")),
                ("LEFTPADDING", (0, 0), (-1, -1), 8),
                ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                ("TOPPADDING", (0, 0), (-1, -1), 7),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
            ]
        )
    )
    story.append(callout_table)

    story.append(p("1. 数据范围", styles["h1"]))
    scope = [
        f"纳入：主表 RD Lifetime test-96ch P1KH 中的 cycle、% Life、CV across all dispenses、plate1-plate5、CV of the worst channel、LINK、Notes、Fail Image、Failed Channels，以及 Test Plan 中的计划与备注。",
        "排除：所有 Pressure 工作表、Pressure Test link 列、外链 Pressure Summary、压力阈值、压力 holding/leak-rate/insert 等结果。",
        f"样本数量：{total} 台 production pipette；状态为 Finish 的样本：{finished} 台；最新记录：{fmt_cycles(latest_cycles)} cycles / {fmt_life(latest_life)}。",
    ]
    for item in scope:
        story.append(p("• " + item, styles["body"]))

    story.append(p("2. 汇总判定（非 Pressure）", styles["h1"]))
    summary_rows = [
        [
            p("Pipette", styles["table_center"]),
            p("状态", styles["table_center"]),
            p("最新 cycles", styles["table_center"]),
            p("最新 life", styles["table_center"]),
            p("最新 CV", styles["table_center"]),
            p("最新 worst CV", styles["table_center"]),
            p("非 Pressure 风险", styles["table_center"]),
            p("主要判断", styles["table_center"]),
        ]
    ]
    for s in samples:
        risk, reason = classify(s)
        latest = s["latest_numeric"]
        summary_rows.append(
            [
                p(s["pipette"], styles["table"]),
                p(s["status"], styles["table_center"]),
                p(fmt_cycles(latest["cycles"]), styles["table_center"]),
                p(fmt_life(latest["life"]), styles["table_center"]),
                p(fmt_num(latest["cv"]), styles["table_center"]),
                p(fmt_num(latest["worst"]), styles["table_center"]),
                p(risk, styles["table_center"]),
                p(reason, styles["table"]),
            ]
        )
    story.append(make_table(summary_rows, [30 * mm, 14 * mm, 23 * mm, 19 * mm, 18 * mm, 22 * mm, 18 * mm, 96 * mm]))

    story.append(p("3. 光学/CV 趋势重点", styles["h1"]))
    trend_rows = [
        [
            p("Pipette", styles["table_center"]),
            p("最大 CV 节点", styles["table_center"]),
            p("最大 CV", styles["table_center"]),
            p("最大 worst-channel CV 节点", styles["table_center"]),
            p("最大 worst-channel CV", styles["table_center"]),
            p("最新口径", styles["table_center"]),
            p("说明", styles["table_center"]),
        ]
    ]
    for s in samples:
        max_cv = s["max_cv"]
        max_worst = s["max_worst"]
        latest = s["latest_numeric"]
        latest_scope = f"{s['latest_plate_count']} plates"
        explanation = []
        if s["latest_plate_count"] < 5:
            explanation.append("1020k 使用 3 盘数据，plate4/5 为 #DIV/0!；与早期 5 盘口径不可直接等价。")
        if max_cv["row"] == latest["row"]:
            explanation.append("最大 CV 出现在最新节点。")
        if max_worst["row"] != max_cv["row"]:
            explanation.append("最大 worst-channel CV 与最大整体 CV 不在同一节点。")
        trend_rows.append(
            [
                p(s["pipette"], styles["table"]),
                p(f"{fmt_cycles(max_cv['cycles'])}<br/>{fmt_life(max_cv['life'])}", styles["table_center"]),
                p(fmt_num(max_cv["cv"]), styles["table_center"]),
                p(f"{fmt_cycles(max_worst['cycles'])}<br/>{fmt_life(max_worst['life'])}", styles["table_center"]),
                p(fmt_num(max_worst["worst"]), styles["table_center"]),
                p(latest_scope, styles["table_center"]),
                p(" ".join(explanation) or "未见额外口径说明。", styles["table"]),
            ]
        )
    story.append(make_table(trend_rows, [30 * mm, 26 * mm, 18 * mm, 31 * mm, 23 * mm, 19 * mm, 98 * mm]))

    story.append(p("4. 异常与备注", styles["h1"]))
    event_rows = [
        [
            p("Pipette", styles["table_center"]),
            p("节点", styles["table_center"]),
            p("类型", styles["table_center"]),
            p("记录内容", styles["table_center"]),
        ]
    ]
    for s in samples:
        for fail in s["explicit_fails"]:
            event_rows.append(
                [
                    p(s["pipette"], styles["table"]),
                    p(f"{fmt_cycles(fail['cycles'])}<br/>{fmt_life(fail['life'])}", styles["table_center"]),
                    p("显式 FAIL", styles["table_center"]),
                    p(fail["note"] or "结果单元格含 fail，但备注为空。", styles["table"]),
                ]
            )
        for note in s["notes"]:
            event_rows.append(
                [
                    p(s["pipette"], styles["table"]),
                    p(f"{fmt_cycles(note['cycles'])}<br/>{fmt_life(note['life'])}", styles["table_center"]),
                    p("备注", styles["table_center"]),
                    p(note["note"] or note["failed_channels"], styles["table"]),
                ]
            )
    if len(event_rows) == 1:
        event_rows.append([p("-", styles["table_center"]), p("-", styles["table_center"]), p("-", styles["table_center"]), p("未记录备注或 fail。", styles["table"])])
    story.append(make_table(event_rows, [34 * mm, 30 * mm, 22 * mm, 175 * mm]))

    story.append(p("5. Test Plan 与数据质量", styles["h1"]))
    start_date = data["start_date"].isoformat() if data["start_date"] else "-"
    finish_date = data["finish_392"].isoformat() if data["finish_392"] else "-"
    quality_points = [
        f"Test Plan 显示起始日期为 {start_date}，392% 计划完成日期为 {finish_date}；主表最新节点已到 392.3%，且状态为 Finish。",
        "Test Plan 中仍保留“继续跑到 fail 为止”等过程备注；如本报告作为结案报告，建议补充最终停止标准和最终审批记录。",
        f"主表 `TOTLE FAIL` 列在 5 台样机首行均为 0，但主表记录了 {explicit_fail_count} 个显式 fail 节点；建议同步修正统计字段。",
        "1020k 最新光学节点为 3 盘测试口径，建议在主表中统一标注，避免和 5 盘历史数据直接比较。",
    ]
    for item in quality_points:
        story.append(p("• " + item, styles["body"]))

    story.append(p("6. 建议", styles["h1"]))
    actions = [
        "优先复测 `P1KHV3620250624A05` 的 1020k 光学数据，确认 CV 6.839 是否为真实退化、三盘口径影响或数据处理问题。",
        "复核 `P1KHV3620250624A04` 的 520k 漏油/背板有油/传感器报错记录，并确认最新高 CV 是否仍与该事件相关。",
        "对 `P1KHV3620241217A05` 的 175% fail 记录做结案说明，重点确认 C5/C7 吸水少是否复现或已排除。",
        "对 `P1KHV3620250624A01` 的 H1/G12 low %D 历史备注保留跟踪，但从最新非 Pressure 数据看，其 CV 已明显回落。",
        "若需要正式 lifetime PASS/FAIL 判定，请补充光学 CV 与吸水检查的明确规格阈值；本报告未使用 Pressure 测试结果参与判定。",
    ]
    for i, item in enumerate(actions, start=1):
        story.append(p(f"{i}. {item}", styles["body"]))

    doc.build(story)


if __name__ == "__main__":
    build_pdf(extract_data())
    print(OUT)
